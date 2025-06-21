import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class AttendanceSaver:
    def __init__(self, sheet_path="attendance_sheet.xlsx"):
        self.sheet_path = sheet_path

    def save(self, new_attendance_df: pd.DataFrame) -> str:
        today = datetime.today().strftime("%Y-%m-%d")

        if os.path.exists(self.sheet_path):
            df = pd.read_excel(self.sheet_path)
        else:
            df = pd.DataFrame(columns=["Name"])

        existing_names = set(df["Name"]) if "Name" in df.columns else set()
        new_names = set(new_attendance_df["Name"])
        missing_names = new_names - existing_names

        if missing_names:
            new_rows = pd.DataFrame({"Name": list(missing_names)})
            df = pd.concat([df, new_rows], ignore_index=True)

        if today not in df.columns:
            df[today] = ""

        for _, row in new_attendance_df.iterrows():
            df.loc[df["Name"] == row["Name"], today] = row["Status"]

        df = df.sort_values("Name").reset_index(drop=True)
        df.to_excel(self.sheet_path, index=False)

        self.apply_colors()

        return os.path.abspath(self.sheet_path)

    def apply_colors(self):
        wb = load_workbook(self.sheet_path)
        ws = wb.active

        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") 
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") 

        for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):
            for cell in row:
                if cell.value == "Present":
                    cell.fill = green
                elif cell.value == "Absent":
                    cell.fill = red
                elif cell.value == "Late":
                    cell.fill = yellow

        wb.save(self.sheet_path)
